from __future__ import annotations

import hashlib
import json
import os
import traceback
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TEST_FILES = Path(os.environ.get("TEST_FILES_DIR", "/test_files"))
WORKSPACE = Path(os.environ.get("WORKSPACE_DIR", "/workspace"))
EVAL_DIR = Path(os.environ.get("EVAL_DIR", "/eval"))
RESULT_PATH = EVAL_DIR / "code_result.json"

INPUT_NAME = "vendor_onboarding_hold_queue.xlsx"
TARGET_SHEET = "HoldQueue"
HEADERS = [
    "Application ID",
    "Canonical Vendor",
    "Office Code",
    "Matched Root",
    "Matched Search Volume",
    "Win Count",
    "Spend Status",
    "Trigger Type",
    "Trigger Key",
]
EXPECTED_ROWS = [
    ["APP-4101", "Apex Components", 310, "export", 80, 1, "KNOWN_SPEND", "risk_code", "EXP"],
    ["APP-4102", "NorthBridge Labs", 120, "medical", 70, 1, "ZERO_SPEND", "risk_code", "MED"],
    ["APP-4103", "Cobalt Mining", 410, "mining", 65, 2, "UNKNOWN_SPEND", "country_category", "ZA|Mining"],
    ["APP-4104", "OldTown Medical", 122, "sanctions", 110, 1, "KNOWN_SPEND", "risk_code_resubmitted", "SAN"],
    ["APP-4107", "Nadir Analytics", 230, "security", 60, 2, "UNKNOWN_SPEND", "risk_code", "SEC"],
]
SCORED_ROWS = 10
SCORED_COLS = 9
SOURCE_SHEETS = [
    "Applications",
    "OfficeMap",
    "SearchTerms",
    "RootWords",
    "Scoreboard",
    "RiskRules",
    "ExistingVendors",
]
PROTECTED_CELLS = {
    "Applications": ["L1", "L2"],
    "OfficeMap": ["F1", "F2"],
    "SearchTerms": ["D1", "D2"],
    "HoldQueue": ["K1", "K2"],
}


def write_result(resolved: bool, score: float, reason: str) -> None:
    EVAL_DIR.mkdir(parents=True, exist_ok=True)
    payload = {"resolved": bool(resolved), "score": float(score), "reason": str(reason)[:2000]}
    RESULT_PATH.write_text(json.dumps(payload, ensure_ascii=False) + "\n", encoding="utf-8")


def norm(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def values_equal(actual: Any, expected: Any) -> bool:
    actual = norm(actual)
    expected = norm(expected)
    if isinstance(expected, (int, float)) and isinstance(actual, str):
        try:
            actual = float(actual)
        except ValueError:
            return False
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return abs(float(actual) - float(expected)) <= 1e-9
    return actual == expected


def snapshot(wb, sheet: str) -> list[list[Any]]:
    return [[norm(cell.value) for cell in row] for row in wb[sheet].iter_rows()]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def candidate_workbooks() -> list[Path]:
    candidates: list[Path] = []
    if WORKSPACE.exists():
        candidates.extend(WORKSPACE.rglob("*.xlsx"))
    if candidates:
        original = (TEST_FILES / INPUT_NAME).resolve()
        filtered = [p for p in candidates if p.resolve() != original]
        candidates = filtered or candidates
        return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    if TEST_FILES.exists():
        return sorted(TEST_FILES.rglob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return []


def formula_cells(ws) -> list[str]:
    cells: list[str] = []
    for row in range(2, 2 + SCORED_ROWS):
        for col in range(1, SCORED_COLS + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.startswith("="):
                cells.append(ws.cell(row=row, column=col).coordinate)
    return cells


def evaluate(candidate: Path) -> tuple[float, str]:
    input_path = TEST_FILES / INPUT_NAME
    if not input_path.exists():
        return 0.0, f"Missing input workbook at {input_path}"
    try:
        if candidate.resolve() == input_path.resolve():
            return 0.0, "Only the original uploaded workbook was found; no model-produced workbook was detected"
    except OSError:
        pass

    try:
        input_wb = load_workbook(input_path, data_only=False)
        candidate_wb = load_workbook(candidate, data_only=False)
    except Exception as exc:
        return 0.0, f"Workbook open failed: {exc}"

    required = [*SOURCE_SHEETS, TARGET_SHEET]
    missing = [sheet for sheet in required if sheet not in candidate_wb.sheetnames]
    if missing:
        return 0.0, f"Missing required sheets: {missing}"

    changed = []
    for sheet in SOURCE_SHEETS:
        if snapshot(input_wb, sheet) != snapshot(candidate_wb, sheet):
            changed.append(sheet)
    for sheet, coords in PROTECTED_CELLS.items():
        for coord in coords:
            if norm(input_wb[sheet][coord].value) != norm(candidate_wb[sheet][coord].value):
                changed.append(f"{sheet}!{coord}")
    if changed:
        return 0.0, f"Source/protected content changed: {sorted(set(changed))}"

    ws = candidate_wb[TARGET_SHEET]
    actual_headers = [norm(ws.cell(row=1, column=col).value) for col in range(1, SCORED_COLS + 1)]
    if actual_headers != HEADERS:
        return 0.0, f"HoldQueue header mismatch: {actual_headers!r}"

    formulas = formula_cells(ws)
    if formulas:
        return 0.0, f"Formulas are not allowed in A2:I11: {formulas}"

    expected_full = EXPECTED_ROWS + [[""] * SCORED_COLS for _ in range(SCORED_ROWS - len(EXPECTED_ROWS))]
    total = SCORED_ROWS * SCORED_COLS
    correct = 0
    mismatches = []
    for r_idx, expected_row in enumerate(expected_full, start=2):
        for c_idx, expected in enumerate(expected_row, start=1):
            actual = ws.cell(row=r_idx, column=c_idx).value
            if values_equal(actual, expected):
                correct += 1
            elif len(mismatches) < 8:
                mismatches.append(
                    f"{ws.cell(row=r_idx, column=c_idx).coordinate} expected {expected!r} found {norm(actual)!r}"
                )

    score = correct / total
    if score == 1.0:
        return 1.0, f"All expected values matched; input sha256={sha256(input_path)}"
    return score, f"{correct}/{total} cells matched. Examples: {mismatches}"


def main() -> None:
    try:
        candidates = candidate_workbooks()
        if not candidates:
            write_result(False, 0.0, "No .xlsx workbook found in /workspace or /test_files")
            return
        candidate = candidates[0]
        score, reason = evaluate(candidate)
        write_result(score >= 0.85, score, f"Checked {candidate}: {reason}")
    except Exception:
        write_result(False, 0.0, traceback.format_exc())


if __name__ == "__main__":
    main()
